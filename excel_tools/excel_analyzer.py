from openpyxl import load_workbook, Workbook
wb=load_workbook(r"C:\Users\Gautam G\OneDrive\Desktop\students.xlsx")
sheet=wb.active
highest=-1
lowest=101
topstudent=""
bottomstudent=""    
total=0
count=0
new_wb=Workbook()
new_sheet=new_wb.active
for row in sheet.iter_rows(min_row=2,values_only=True):
    name,marks=row
    if name is None or marks is None:
        continue    
    if marks>highest:
        highest=marks
        topstudent=name
    if marks<lowest:
        lowest=marks
        bottomstudent=name
    total+=marks
    count+=1
    
average=total/count
print(f"The highest marks are {highest} obtained by {topstudent}.") 
print(f"The lowest marks are {lowest} obtained by {bottomstudent}.")
print(f"The average marks of the students are {average:.2f}.")
print("Total students:", count)
new_sheet.append(["Name", "Marks", "Result"])
for row in sheet.iter_rows(min_row=2,values_only=True):
    name,marks=row
    if name is None or marks is None:
        continue
    if marks>=50:
        result="Pass"
    else:
        result="Fail"
    new_sheet.append([name, marks, result])

new_sheet.append([])
new_sheet.append(["Summary"])
new_sheet.append(["Highest Marks", highest, f"Obtained by {topstudent}"])
new_sheet.append(["Lowest Marks", lowest, f"Obtained by {bottomstudent}"])
new_sheet.append(["Total Students", count])
new_sheet.append(["Average Marks", f"{average:.2f}"])   
new_wb.save(r"C:\Users\Gautam G\OneDrive\Desktop\student_report.xlsx")
