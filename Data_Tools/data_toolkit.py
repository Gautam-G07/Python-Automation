import csv
import os
from openpyxl import  load_workbook
def clean_csv(input_file):
    if not os.path.exists(input_file):
        print("File not found. Please check the path and try again.")
        return []
    clean_data=[]
    with open(input_file, "r") as file:
        reader = csv.reader(file)
        next(reader,None) 
        for row in reader:
            if len(row) != 2:
                continue
            name,marks=row
            name=name.strip()
            try:
                marks=int(marks.strip())
            except ValueError:
                continue
            if name=="":
                continue
            clean_data.append([name, marks])
    return clean_data
def save_csv(clean_data, output_file):           

    with open(output_file, "w", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(["Name", "Marks"])
        writer.writerows(clean_data)
    print("Clean CSV FILE CREATED!")
    
def analyze_csv(clean_data):
        if not clean_data:
            print("No data to analyze.")
            return
        highest=-1
        topstudent=""    
        for name, marks in clean_data:
            if marks>highest:
                highest=marks
                topstudent=name
        print(f"The highest marks are {highest} obtained by {topstudent}.")
def analyze_excel(file):
    if not os.path.exists(file):
        print("File not found. Please check the path and try again.")
        return
    workbook = load_workbook(file)
    sheet = workbook.active
    highest=-1
    topstudent=""
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, marks = row
        if name is None or marks is None:
            continue
        if  marks > highest:
            highest = marks
            topstudent = name
    print(f"The highest marks are {highest} obtained by {topstudent}.")
while True:
    print("\n-------DATA TOOLKIT------")
    print("1. Clean CSV")
    print("2. Analyze CSV")
    print("3. Analyze Excel")
    print("4. Exit")
    choice = input("Enter your choice: ")
    if choice == "1":
        input_file = input("Enter the input CSV file path: ")
        output_file = input("Enter the output CSV file path: ")
        data=clean_csv(input_file)
        save_csv(data, output_file)
    elif choice == "2":
        input_file = input("Enter the input CSV file path: ")
        data=clean_csv(input_file)
        analyze_csv(data)
    elif choice == "3":
        excel_file = input("Enter the Excel file path: ")
        analyze_excel(excel_file)
    elif choice == "4":
        break
    else:
        print("Invalid choice. Please try again.")